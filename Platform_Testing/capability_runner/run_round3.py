"""
Round 3 — Data Volume Runner

Tests rule adherence at fixed passing tiers across increasing payload sizes.
Only tests tiers where each model passed cleanly in Rounds 1 & 2.

Runs:
  Opus   — Scope T1-T5, Chained T1-T5, Format T1-T5  (15 per payload)
  Sonnet — Chained T1-T3, Format T1-T3               (6 per payload)
  Haiku  — Format T1-T2                               (2 per payload)
  Payloads: Medium (40 rows), Large (80 rows), XL (120 rows)
  Total: 69 runs x 2 API calls = 138 API calls

Usage:
    python run_round3.py
    python run_round3.py --dry-run
    python run_round3.py --models claude-opus-4-8
    python run_round3.py --payloads Medium Large
    python run_round3.py --rule-sets Scope

Requirements:
    pip install anthropic openpyxl   (or swap provider in connector.py)
    Set API_KEY below or in your environment.
"""

import argparse
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(__file__))

from connector import create_client, call_model
from shared import (
    judge_response, build_user_message, build_system_prompt,
    detect_collapse, load_financial_data, soft_pct,
    PASS_FILL, FAIL_FILL, EVEN_FILL, ODD_FILL,
    _border, _font, _align,
)

# ── Config ────────────────────────────────────────────────────────────

WORKBOOK      = os.path.join(os.path.dirname(__file__), "..", "capability_testing_round3.xlsx")
RESULTS_SHEET = "Round 3 Results"

API_KEY = ""

DEFAULT_MODELS   = ["claude-opus-4-8", "claude-sonnet-4-6", "claude-haiku-4-5"]
DEFAULT_PAYLOADS = ["Medium", "Large", "XL"]

# ── Rule sets ─────────────────────────────────────────────────────────

SCO_RULES = [
    "Only reference companies present in the attached dataset.",
    "Only use figures that appear explicitly in the attached data. Do not calculate or derive new figures.",
    "Do not reference any news, events, or market context not in the dataset.",
    "Do not describe any trend that spans periods not covered in the dataset.",
    "Do not compare any figure to a historical average unless that average is in the dataset.",
]
CHN_RULES = [
    "If diff_pct exceeds 5% AND netShChg is negative, flag both in the same sentence.",
    "If actual EPS missed AND surprise_percent is below -5%, label it a significant miss.",
    "If actual EPS beat AND netShChg is positive, note alignment between earnings and insider sentiment.",
    "If actual EPS beat AND surprise_percent exceeds +10%, label it an outsized beat.",
    "If diff_pct exceeds 5% AND actual EPS missed, open the ticker section with a risk note.",
]
FMT_RULES = [
    "Begin every response with the exact phrase: \"Here is what the data shows:\"",
    "End every response with the exact phrase: \"Data sourced from your portfolio dataset.\"",
    "Write the company full name followed by the ticker in parentheses, e.g., Apple Inc. (AAPL).",
    "All dollar amounts must be formatted to exactly two decimal places (e.g., $1.23, not $1.2 or $1.234).",
    "All percentage values must include an explicit + or - sign (e.g., +4.50%, -1.20%).",
]

# What each model tests: {model: [(rule_set_name, rules, max_tier)]}
MODEL_PLAN = {
    "claude-opus-4-8": [
        ("Scope",   SCO_RULES, 5),
        ("Chained", CHN_RULES, 5),
        ("Format",  FMT_RULES, 5),
    ],
    "claude-sonnet-4-6": [
        ("Chained", CHN_RULES, 3),
        ("Format",  FMT_RULES, 3),
    ],
    "claude-haiku-4-5": [
        ("Format",  FMT_RULES, 2),
    ],
}

PAYLOAD_SHEETS = {
    "Medium": "Data - Medium",
    "Large":  "Data - Large",
    "XL":     "Data - XL",
}
PAYLOAD_ROWS = {"Medium": 40, "Large": 80, "XL": 120}


# ── Validation ────────────────────────────────────────────────────────

def validate_plan(wb):
    print("\nValidating test plan...")
    errors = []
    for model, rule_sets in MODEL_PLAN.items():
        for rule_set_name, rules, max_tier in rule_sets:
            if max_tier > len(rules):
                errors.append(
                    f"{model} / {rule_set_name}: max_tier {max_tier} exceeds "
                    f"available rules ({len(rules)})"
                )
    for payload, sheet_name in PAYLOAD_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            errors.append(f"Sheet '{sheet_name}' not found in workbook")
    if errors:
        for e in errors:
            print(f"  ERROR: {e}")
        raise SystemExit("Validation failed.")
    print("  Validation passed.\n")


# ── Results writing ───────────────────────────────────────────────────

def find_next_row(ws):
    row = 4
    while ws.cell(row, 9).value not in (None, ""):
        row += 1
    return row


def write_row(ws, row_num, data_row, is_pass):
    bg  = PASS_FILL if is_pass else FAIL_FILL
    alt = EVEN_FILL if row_num % 2 == 0 else ODD_FILL
    center_cols = {1, 2, 3, 4, 5, 6, 9, 10, 11, 12}

    for col_idx, value in enumerate(data_row, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font      = _font()
        cell.border    = _border()
        cell.alignment = _align("center" if col_idx in center_cols else "left")
        if col_idx == 9:
            cell.fill = bg
            cell.font = _font(bold=True)
        elif col_idx == 10:
            cell.fill = bg
        else:
            cell.fill = alt
    ws.row_dimensions[row_num].height = 40


def backfill_first_failure(ws, model, rule_set, tier):
    first_fail = None
    for r in range(4, ws.max_row + 1):
        if (ws.cell(r, 1).value == model and
                ws.cell(r, 4).value == rule_set and
                ws.cell(r, 5).value == tier):
            if ws.cell(r, 9).value == "FAIL" and first_fail is None:
                first_fail = ws.cell(r, 2).value
            ws.cell(r, 11).value = first_fail if first_fail else "No failure"


# ── Main ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Round 3 — Data Volume Runner")
    parser.add_argument("--workbook",    default=WORKBOOK)
    parser.add_argument("--dry-run",     action="store_true")
    parser.add_argument("--models",      nargs="+", default=DEFAULT_MODELS)
    parser.add_argument("--payloads",    nargs="+", default=DEFAULT_PAYLOADS)
    parser.add_argument("--rule-sets",   nargs="+", default=None)
    parser.add_argument("--judge-model", default="claude-opus-4-8")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.workbook)
    validate_plan(wb)
    ws = wb[RESULTS_SHEET]

    client = None
    if not args.dry_run:
        api_key = API_KEY or os.environ.get("ANTHROPIC_API_KEY", "")
        if not api_key:
            raise SystemExit("API key not set. Add it to API_KEY in run_round3.py or set ANTHROPIC_API_KEY.")
        client = create_client(api_key)

    total_runs = sum(
        len(args.payloads) * max_tier
        for model in args.models
        if model in MODEL_PLAN
        for rule_set_name, rules, max_tier in MODEL_PLAN[model]
        if not args.rule_sets or rule_set_name in args.rule_sets
    )

    print(f"\nRound 3 — Data Volume Testing")
    print(f"Models   : {', '.join(args.models)}")
    print(f"Payloads : {', '.join(args.payloads)}")
    print(f"Runs     : {total_runs}  ({total_runs * 2} API calls with judge)")
    print(f"Workbook : {args.workbook}\n")

    for model in args.models:
        if model not in MODEL_PLAN:
            print(f"  Skipping {model} — not in MODEL_PLAN")
            continue

        print(f"\n{'='*60}")
        print(f"MODEL: {model}")
        print(f"{'='*60}")

        for rule_set_name, rules, max_tier in MODEL_PLAN[model]:
            if args.rule_sets and rule_set_name not in args.rule_sets:
                continue

            print(f"\n  Rule Set: {rule_set_name}  (testing T1-T{max_tier})")

            for tier in range(1, max_tier + 1):
                cumulative        = rules[:tier]
                active_rules_text = "\n".join(f"{i+1}. {r}" for i, r in enumerate(cumulative))
                system            = build_system_prompt(rule_set_name, cumulative)
                payload_results   = []

                for payload in args.payloads:
                    sheet_name     = PAYLOAD_SHEETS[payload]
                    row_count      = PAYLOAD_ROWS[payload]
                    financial_data = load_financial_data(wb, sheet_name, header_row=2)
                    user_msg       = build_user_message(financial_data)

                    print(f"    T{tier} | {payload:6s} ({row_count} rows)", end="", flush=True)

                    if args.dry_run:
                        print(" [DRY RUN]")
                        continue

                    print(" → calling...", end="", flush=True)
                    response_text = call_model(client, system, user_msg, model)

                    print(" → judging...", end="", flush=True)
                    verdict    = judge_response(client, cumulative, response_text, args.judge_model)

                    passed_n   = verdict["passed"]
                    total_r    = verdict["total"]
                    violations = verdict["violations"]
                    pct        = soft_pct(passed_n, total_r)
                    pass_fail  = "PASS" if not violations else "FAIL"
                    is_pass    = pass_fail == "PASS"

                    payload_results.append((payload, passed_n, total_r))
                    collapse = detect_collapse(payload_results)

                    print(f" → {pass_fail} {pct}")

                    row_num  = find_next_row(ws)
                    data_row = [
                        model,                                       # col 1  Model
                        payload,                                     # col 2  Payload Size
                        row_count,                                   # col 3  Row Count
                        rule_set_name,                               # col 4  Rule Set
                        tier,                                        # col 5  Tier
                        len(cumulative),                             # col 6  Rule Count
                        active_rules_text,                           # col 7  Active Rules
                        "; ".join(violations) if violations else "", # col 8  Rules Violated
                        pass_fail,                                   # col 9  Pass / Fail
                        pct,                                         # col 10 Soft Score
                        "",                                          # col 11 First Failure Payload (backfilled)
                        collapse,                                    # col 12 Collapse
                        "",                                          # col 13 Notes
                    ]
                    write_row(ws, row_num, data_row, is_pass)

                if not args.dry_run:
                    backfill_first_failure(ws, model, rule_set_name, tier)

        if not args.dry_run:
            wb.save(args.workbook)
            print(f"\n  Saved after {model}")

    if not args.dry_run:
        wb.save(args.workbook)
        print(f"\nDone. Results saved to {args.workbook}")
    else:
        print("\n[DRY RUN] No API calls made. No file saved.")


if __name__ == "__main__":
    main()
