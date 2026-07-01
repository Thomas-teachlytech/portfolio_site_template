"""
Round 2 — Cross-Category Combination Runner

Reads combination rules from capability_testing_round2.xlsx, calls the configured
model for each combo x tier x model, judges each response, and writes results
to the 'Round 2 Results' sheet.

Usage:
    python run_round2.py                                    # all combos, all tiers, all models
    python run_round2.py --dry-run
    python run_round2.py --models claude-opus-4-8
    python run_round2.py --combo M1
    python run_round2.py --combo M1 M4 --tiers 1 2 3

Requirements:
    pip install anthropic openpyxl   (or swap provider in connector.py)
    Set API_KEY below or in your environment.
"""

import argparse
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(__file__))

from connector import create_client, call_model
from shared import (
    judge_response, build_user_message, load_financial_data, soft_pct,
    detect_collapse, PASS_FILL, FAIL_FILL, EVEN_FILL, ODD_FILL,
    _border, _font, _align,
)

# ── Config ────────────────────────────────────────────────────────────

WORKBOOK      = os.path.join(os.path.dirname(__file__), "..", "capability_testing_round2.xlsx")
DESIGN_SHEET  = "Combination Design"
RESULTS_SHEET = "Round 2 Results"
DATA_SHEET    = "Round 1 Data"

API_KEY = ""

DEFAULT_MODELS = [
    "claude-opus-4-8",
    "claude-sonnet-4-6",
    "claude-haiku-4-5",
]

# Combination Design sheet layout
DESIGN_DATA_START_ROW = 5
COMBO_ID_COL       = 1
COMBO_LABEL_COL    = 2
COMBO_CATS_COL     = 3
COMBO_WEIGHT_COL   = 4
COMBO_BAND_COL     = 5
COMBO_INTERACT_COL = 6
COMBO_QUESTION_COL = 7
TIER_START_COL     = 8   # T1=col8 … T7=col14

RESULTS_DATA_START_ROW = 4


# ── Data loading ──────────────────────────────────────────────────────

def load_combinations(wb):
    ws = wb[DESIGN_SHEET]
    combos = {}
    for row in ws.iter_rows(min_row=DESIGN_DATA_START_ROW, values_only=True):
        combo_id = row[COMBO_ID_COL - 1]
        if not combo_id:
            continue
        tier_rules = []
        for t in range(7):
            val = row[TIER_START_COL - 1 + t]
            tier_rules.append(str(val).strip() if val else "")
        combos[combo_id] = {
            "label":       row[COMBO_LABEL_COL - 1],
            "categories":  row[COMBO_CATS_COL - 1],
            "weight":      row[COMBO_WEIGHT_COL - 1],
            "band":        row[COMBO_BAND_COL - 1],
            "interaction": row[COMBO_INTERACT_COL - 1],
            "question":    row[COMBO_QUESTION_COL - 1],
            "tiers":       tier_rules,
        }
    return combos


# ── Validation ────────────────────────────────────────────────────────

def parse_rules(tier_rule_text: str) -> list[str]:
    return [r.strip() for r in tier_rule_text.split(" | ") if r.strip()]


def validate_combinations(combos):
    print("\nValidating combination rules...")
    problems = []

    for combo_id, info in combos.items():
        tiers = info["tiers"]
        for t_idx, rule_text in enumerate(tiers):
            tier_num = t_idx + 1
            location = f"{combo_id} T{tier_num}"
            if not rule_text:
                problems.append(f"{location}: empty")
                continue
            rules = parse_rules(rule_text)
            if len(rules) < tier_num:
                problems.append(f"{location}: expected >= {tier_num} rules, found {len(rules)}")
            if t_idx > 0:
                prev = parse_rules(tiers[t_idx - 1])
                if len(rules) <= len(prev):
                    problems.append(f"{location}: rule count ({len(rules)}) not > T{tier_num-1} ({len(prev)})")
            for rule in rules:
                if len(rule) < 10:
                    problems.append(f"{location}: suspiciously short rule — '{rule}'")

        rule_counts = [len(parse_rules(t)) for t in tiers if t]
        status = "OK" if not any(combo_id in p for p in problems) else "PROBLEM"
        print(f"  {combo_id:4}  {info['band']:8}  rules per tier: {rule_counts}  [{status}]")

    if problems:
        print(f"\n  {len(problems)} problem(s) found:")
        for p in problems:
            print(f"    - {p}")
        return False

    print("\n  All combinations valid.\n")
    return True


# ── Prompt building ───────────────────────────────────────────────────

def build_system_prompt(combo_id, tier_rule_text):
    rules = parse_rules(tier_rule_text)
    numbered = "\n".join(f"{i+1}. {r}" for i, r in enumerate(rules))
    return (
        f"You are a financial data summarisation assistant (combination {combo_id}).\n"
        f"Follow ALL of the rules below exactly. There are {len(rules)} active rules.\n\n"
        f"RULES:\n{numbered}"
    )


# ── Results writing ───────────────────────────────────────────────────

def find_next_row(ws):
    row = RESULTS_DATA_START_ROW
    while ws.cell(row, 10).value not in (None, ""):
        row += 1
    return row


def write_row(ws, row_num, data_row, is_pass):
    bg  = PASS_FILL if is_pass else FAIL_FILL
    alt = EVEN_FILL if row_num % 2 == 0 else ODD_FILL
    center_cols = {1, 2, 4, 5, 6, 7, 10, 11, 12, 13}

    for col_idx, value in enumerate(data_row, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font      = _font(bold=(col_idx == 10))
        cell.border    = _border()
        cell.alignment = _align("center" if col_idx in center_cols else "left")
        if col_idx in (10, 11):
            cell.fill = bg
        else:
            cell.fill = alt
    ws.row_dimensions[row_num].height = 45


def backfill_first_failure(ws, combo_id, model):
    first_fail = None
    for r in range(RESULTS_DATA_START_ROW, ws.max_row + 1):
        if ws.cell(r, 1).value == model and ws.cell(r, 2).value == combo_id:
            if ws.cell(r, 10).value == "FAIL" and first_fail is None:
                first_fail = ws.cell(r, 6).value
            ws.cell(r, 12).value = first_fail if first_fail else "No failure"


# ── Main ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Round 2 — Cross-Category Combinations")
    parser.add_argument("--workbook",    default=WORKBOOK)
    parser.add_argument("--dry-run",     action="store_true")
    parser.add_argument("--models",      nargs="+", default=DEFAULT_MODELS)
    parser.add_argument("--judge-model", default="claude-opus-4-8")
    parser.add_argument("--combo",       nargs="+", help="Combination IDs e.g. H1 M1 M4")
    parser.add_argument("--tiers",       nargs="+", type=int)
    args = parser.parse_args()

    wb       = openpyxl.load_workbook(args.workbook)
    combos   = load_combinations(wb)
    fin_data = load_financial_data(wb, DATA_SHEET, header_row=1)
    ws       = wb[RESULTS_SHEET]
    user_msg = build_user_message(fin_data)

    if not validate_combinations(combos):
        raise SystemExit("Validation failed — fix the Combination Design sheet before running.")

    client = None
    if not args.dry_run:
        api_key = API_KEY or os.environ.get("ANTHROPIC_API_KEY", "")
        if not api_key:
            raise SystemExit("API key not set. Add it to API_KEY in run_round2.py or set ANTHROPIC_API_KEY.")
        client = create_client(api_key)

    combo_ids = args.combo if args.combo else sorted(combos.keys())
    tiers     = args.tiers or list(range(1, 8))
    models    = args.models

    total_runs = len(models) * len(combo_ids) * len(tiers)
    print(f"\nRound 2 — Cross-Category Combinations")
    print(f"Models  : {', '.join(models)}")
    print(f"Combos  : {', '.join(combo_ids)}")
    print(f"Tiers   : {tiers}")
    print(f"Runs    : {total_runs}  |  API calls (with judge): {total_runs * 2}")
    print(f"Workbook: {args.workbook}\n")

    for model in models:
        print(f"\n{'='*60}")
        print(f"MODEL: {model}")
        print(f"{'='*60}")

        for combo_id in combo_ids:
            if combo_id not in combos:
                print(f"  WARNING: '{combo_id}' not found — skipping")
                continue

            info = combos[combo_id]
            print(f"\n  Combo: {combo_id}  ({info['label']})  Band: {info['band']}")
            tier_results = []

            for tier in tiers:
                tier_idx       = tier - 1
                tier_rule_text = info["tiers"][tier_idx] if tier_idx < len(info["tiers"]) else ""

                if not tier_rule_text:
                    print(f"    T{tier}: no rules — skipping")
                    continue

                rules_list = parse_rules(tier_rule_text)
                system     = build_system_prompt(combo_id, tier_rule_text)

                print(f"    T{tier} ({len(rules_list)} rule(s))", end="", flush=True)

                if args.dry_run:
                    print(" [DRY RUN]")
                    print(f"      Rules: {' | '.join(rules_list)}")
                    continue

                print(" → calling...", end="", flush=True)
                response_text = call_model(client, system, user_msg, model)

                print(" → judging...", end="", flush=True)
                verdict    = judge_response(client, rules_list, response_text, args.judge_model)

                passed_n   = verdict["passed"]
                total_r    = verdict["total"]
                violations = verdict["violations"]
                pct        = soft_pct(passed_n, total_r)
                pass_fail  = "PASS" if not violations else "FAIL"
                is_pass    = pass_fail == "PASS"

                tier_results.append((tier, passed_n, total_r))
                collapse = detect_collapse(tier_results)

                print(f" → {pass_fail} {pct}")

                row_num  = find_next_row(ws)
                data_row = [
                    model,                                        # col 1  Model
                    combo_id,                                     # col 2  Combo ID
                    info["label"],                                # col 3  Combination
                    info["categories"],                           # col 4  Categories
                    info["band"],                                 # col 5  Band
                    tier,                                         # col 6  Tier
                    len(rules_list),                              # col 7  Rule Count
                    tier_rule_text,                               # col 8  Active Rules
                    "; ".join(violations) if violations else "",  # col 9  Rules Violated
                    pass_fail,                                    # col 10 Pass / Fail
                    pct,                                          # col 11 Soft Score
                    "",                                           # col 12 First Failure Tier (backfilled)
                    collapse,                                     # col 13 Collapse
                    "",                                           # col 14 Notes
                ]
                write_row(ws, row_num, data_row, is_pass)

            if not args.dry_run:
                backfill_first_failure(ws, combo_id, model)

    if not args.dry_run:
        wb.save(args.workbook)
        print(f"\nResults saved to {args.workbook}")
    else:
        print("\n[DRY RUN] No API calls made. No file saved.")


if __name__ == "__main__":
    main()
