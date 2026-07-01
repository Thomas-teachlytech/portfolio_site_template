"""
shared.py — utilities shared across all round runners.

Includes: judge logic, prompt helpers, Excel styles, collapse detection.
"""

import textwrap
from datetime import datetime

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from connector import call_model


# ── Judge prompt ──────────────────────────────────────────────────────

JUDGE_SYSTEM = textwrap.dedent("""\
    You are a strict rule-adherence evaluator.
    You will be given:
    1. A set of ACTIVE RULES the AI was supposed to follow.
    2. The AI RESPONSE to a financial-data summarisation task.

    For each active rule output exactly one line:
        PASS: <rule text>   — response clearly satisfies the rule
        FAIL: <rule text>   — response violates or ignores the rule

    After all rule lines output exactly:
        SOFT_SCORE: X/N
    where X = rules passed, N = total active rules.

    Be binary. No partial credit. Evaluate only the rules listed.
""")

USER_PROMPT = (
    "Summarise the earnings performance, insider activity, and recent price movement "
    "for the companies in the attached data."
)


# ── Prompt helpers ────────────────────────────────────────────────────

def build_user_message(financial_data: str) -> str:
    return f"{USER_PROMPT}\n\n<data>\n{financial_data}\n</data>"


def build_system_prompt(label: str, cumulative_rules: list[str]) -> str:
    numbered = "\n".join(f"{i+1}. {r}" for i, r in enumerate(cumulative_rules))
    return textwrap.dedent(f"""\
        You are a financial data summarisation assistant ({label}).
        Follow ALL of the rules below exactly. There are {len(cumulative_rules)} active rules.

        RULES:
        {numbered}
    """).strip()


# ── Judge ─────────────────────────────────────────────────────────────

def judge_response(client, rules: list[str], response_text: str, judge_model: str) -> dict:
    """
    Returns {"passed": int, "total": int, "violations": list[str]}.
    """
    rules_block = "\n".join(f"{i+1}. {r}" for i, r in enumerate(rules))
    user_msg = f"ACTIVE RULES:\n{rules_block}\n\nAI RESPONSE:\n{response_text}"
    raw = call_model(client, JUDGE_SYSTEM, user_msg, judge_model)

    violations, passed = [], 0
    total = len(rules)

    for line in raw.splitlines():
        line = line.strip()
        if line.startswith("FAIL:"):
            violations.append(line[5:].strip())
        elif line.startswith("PASS:"):
            passed += 1

    for line in raw.splitlines():
        if line.startswith("SOFT_SCORE:"):
            try:
                parts = line.split(":")[1].strip().split("/")
                passed = int(parts[0])
                total  = int(parts[1])
            except Exception:
                pass
            break

    return {"passed": passed, "total": total, "violations": violations}


# ── Scoring helpers ───────────────────────────────────────────────────

def soft_pct(passed: int, total: int) -> str:
    return f"{round(100 * passed / total)}%" if total > 0 else "N/A"


def detect_collapse(tier_results: list[tuple]) -> str:
    """Return 'Y' if soft score rate drops vs any previous entry."""
    rates = [p / t if t > 0 else 1.0 for _, p, t in tier_results]
    for i in range(1, len(rates)):
        if rates[i] < rates[i - 1]:
            return "Y"
    return "N"


# ── Financial data loader ─────────────────────────────────────────────

def load_financial_data(wb, sheet_name: str, header_row: int = 1) -> str:
    """
    Read a financial data sheet into a tab-separated string.

    Args:
        wb:         openpyxl workbook
        sheet_name: name of the sheet to read
        header_row: row index (1-based) that contains column headers
    """
    ws = wb[sheet_name]
    headers = [ws.cell(header_row, c).value for c in range(1, 13) if ws.cell(header_row, c).value]
    lines = ["\t".join(str(h) for h in headers)]
    for row in range(header_row + 1, 200):
        ticker = ws.cell(row, 1).value
        if not ticker:
            break
        vals = []
        for c in range(1, 13):
            v = ws.cell(row, c).value
            if isinstance(v, datetime):
                v = v.strftime("%Y-%m-%d")
            vals.append("" if v is None else str(v))
        lines.append("\t".join(vals))
    return "\n".join(lines)


# ── Excel styles ──────────────────────────────────────────────────────

PASS_FILL = PatternFill("solid", fgColor="E2EFDA")
FAIL_FILL = PatternFill("solid", fgColor="FCE4D6")
EVEN_FILL = PatternFill("solid", fgColor="D9E8F5")
ODD_FILL  = PatternFill("solid", fgColor="FFFFFF")


def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _font(bold=False, size=10, color="000000"):
    return Font(name="Arial", bold=bold, size=size, color=color)


def _align(horizontal="left"):
    return Alignment(wrap_text=True, vertical="top", horizontal=horizontal)


def style_cell(cell, col_idx: int, row_num: int, center_cols: set,
               pass_col: int = None, score_col: int = None, is_pass: bool = True):
    """Apply standard formatting to a result cell."""
    alt = EVEN_FILL if row_num % 2 == 0 else ODD_FILL
    cell.border    = _border()
    cell.alignment = _align("center" if col_idx in center_cols else "left")

    if col_idx == pass_col:
        cell.fill = PASS_FILL if is_pass else FAIL_FILL
        cell.font = _font(bold=True)
    elif col_idx == score_col:
        cell.fill = PASS_FILL if is_pass else FAIL_FILL
        cell.font = _font()
    else:
        cell.fill = alt
        cell.font = _font()
