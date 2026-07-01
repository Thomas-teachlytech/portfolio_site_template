"""
Round 1 — Instruction Count Runner

Reads agent rules from capability_testing.xlsx, calls the configured model
for each agent x tier x model combination, judges each response, and writes
results to the 'Results — Instruction Count' sheet.

Usage:
    python run_round1.py                                    # all agents, all tiers, default models
    python run_round1.py --dry-run                          # preview prompts, no API calls
    python run_round1.py --models claude-opus-4-8           # single model
    python run_round1.py --models claude-opus-4-8 claude-sonnet-4-6 claude-haiku-4-5
    python run_round1.py --agent Cnt --tiers 1 2 3          # single agent, specific tiers

Requirements:
    pip install anthropic openpyxl   (or swap provider in connector.py)
    Set API_KEY below or in your environment.
"""

import argparse
import os
import sys

import openpyxl

# Allow running from this folder directly
sys.path.insert(0, os.path.dirname(__file__))

from connector import create_client, call_model
from shared import (
    judge_response, build_user_message, build_system_prompt,
    detect_collapse, load_financial_data, soft_pct,
    PASS_FILL, FAIL_FILL, EVEN_FILL, ODD_FILL,
    _border, _font, _align,
)

# ── Config ────────────────────────────────────────────────────────────

WORKBOOK = os.path.join(os.path.dirname(__file__), "..", "capability_testing.xlsx")
RESULTS_SHEET = "Results — Instruction Count"

API_KEY = ""  # or set via environment variable ANTHROPIC_API_KEY

DEFAULT_MODELS = [
    "claude-opus-4-8",
    "claude-sonnet-4-6",
    "claude-haiku-4-5",
]

# Row in 'Agent Design' sheet where each agent's rules live
AGENT_ROWS = {
    "Cnt": 4,   # Content
    "Fmt": 6,   # Format
    "Sty": 8,   # Style
    "Sit": 10,  # Situation
    "Exa": 12,  # Example
    "Neg": 14,  # Negation
    "Sco": 16,  # Scope
    "Pre": 18,  # Precision
    "Per": 20,  # Persona
    "Chn": 22,  # Chained
    "Pri": 24,  # Priority
}
TIER_COL_OFFSET = 3  # T1 is at col 4 (Agent ID=1, Desc=2, HowToCheck=3, T1=4)


# ── Data loading ──────────────────────────────────────────────────────

def load_agents(wb):
    ws = wb["Agent Design"]
    agents = {}
    for agent_id, row in AGENT_ROWS.items():
        task_type = ws.cell(row, 2).value or ""
        rules = []
        for t in range(7):
            col  = TIER_COL_OFFSET + 1 + t
            rule = ws.cell(row, col).value or ""
            rules.append(str(rule).strip())
        agents[agent_id] = {"task_type": task_type, "rules": rules}
    return agents


# ── Results writing ───────────────────────────────────────────────────

def find_next_row(ws):
    row = 4
    while ws.cell(row, 1).value not in (None, ""):
        row += 1
    return row


def write_row(ws, row_num, data_row, is_pass):
    bg  = PASS_FILL if is_pass else FAIL_FILL
    alt = EVEN_FILL if row_num % 2 == 0 else ODD_FILL
    center_cols = {1, 4, 5, 6, 9, 10, 11, 12}

    for col_idx, value in enumerate(data_row, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font      = _font()
        cell.border    = _border()
        cell.alignment = _align("center" if col_idx in center_cols else "left")
        if col_idx == 9:
            cell.fill = bg
            cell.font = _font(bold=True)
        elif col_idx == 10:
            cell.fill = bg
        else:
            cell.fill = alt
    ws.row_dimensions[row_num].height = 40


def backfill_first_failure(ws, agent_id, model):
    first_fail = None
    for r in range(4, ws.max_row + 1):
        if ws.cell(r, 1).value == model and ws.cell(r, 2).value == agent_id:
            if ws.cell(r, 9).value == "FAIL" and first_fail is None:
                first_fail = ws.cell(r, 5).value
            ws.cell(r, 11).value = first_fail if first_fail else "No failure"


# ── Main ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Round 1 — Instruction Count")
    parser.add_argument("--workbook",    default=WORKBOOK)
    parser.add_argument("--dry-run",     action="store_true")
    parser.add_argument("--models",      nargs="+", default=DEFAULT_MODELS)
    parser.add_argument("--judge-model", default="claude-opus-4-8")
    parser.add_argument("--agent",       help="Single agent ID e.g. Cnt")
    parser.add_argument("--tiers",       nargs="+", type=int)
    args = parser.parse_args()

    wb       = openpyxl.load_workbook(args.workbook)
    agents   = load_agents(wb)
    fin_data = load_financial_data(wb, "Round 1 Data", header_row=1)
    ws       = wb[RESULTS_SHEET]
    user_msg = build_user_message(fin_data)

    client = None
    if not args.dry_run:
        api_key = API_KEY or os.environ.get("ANTHROPIC_API_KEY", "")
        if not api_key:
            raise SystemExit("API key not set. Add it to API_KEY in run_round1.py or set ANTHROPIC_API_KEY.")
        client = create_client(api_key)

    agent_ids = [args.agent] if args.agent else sorted(AGENT_ROWS.keys())
    tiers     = args.tiers or list(range(1, 8))
    models    = args.models

    total_calls = len(models) * len(agent_ids) * len(tiers)
    print(f"\nRound 1 — Instruction Count")
    print(f"Models : {', '.join(models)}")
    print(f"Agents : {', '.join(agent_ids)}")
    print(f"Tiers  : {tiers}")
    print(f"Total  : {total_calls} runs ({total_calls * 2} API calls with judge)")
    print(f"Workbook: {args.workbook}\n")

    for model in models:
        print(f"\n{'='*60}")
        print(f"MODEL: {model}")
        print(f"{'='*60}")

        for agent_id in agent_ids:
            info = agents[agent_id]
            print(f"\n  Agent: {agent_id}")
            tier_results = []

            for tier in tiers:
                cumulative = [r for r in info["rules"][:tier] if r]
                if not cumulative:
                    print(f"    T{tier}: no rules — skipping")
                    continue

                active_rules_text = "\n".join(f"{i+1}. {r}" for i, r in enumerate(cumulative))
                system = build_system_prompt(agent_id, cumulative)

                print(f"    T{tier} ({len(cumulative)} rule(s))", end="", flush=True)

                if args.dry_run:
                    print(" [DRY RUN]")
                    print(f"      SYSTEM (first 200):\n      {system[:200]}...")
                    continue

                print(" → calling...", end="", flush=True)
                response_text = call_model(client, system, user_msg, model)

                print(" → judging...", end="", flush=True)
                verdict = judge_response(client, cumulative, response_text, args.judge_model)

                passed_n   = verdict["passed"]
                total_r    = verdict["total"]
                violations = verdict["violations"]
                pct        = soft_pct(passed_n, total_r)
                pass_fail  = "PASS" if not violations else "FAIL"
                is_pass    = pass_fail == "PASS"

                tier_results.append((tier, passed_n, total_r))
                collapse = detect_collapse(tier_results)

                print(f" → {pass_fail} {pct}")

                row_num  = find_next_row(ws)
                data_row = [
                    model,                                       # col 1  Model
                    agent_id,                                    # col 2  Agent ID
                    agent_id.split("-")[0],                      # col 3  Category code
                    info["task_type"],                           # col 4  Task Type
                    tier,                                        # col 5  Tier
                    len(cumulative),                             # col 6  Rule Count
                    active_rules_text,                           # col 7  Active Rules
                    "; ".join(violations) if violations else "", # col 8  Rules Violated
                    pass_fail,                                   # col 9  Pass / Fail
                    pct,                                         # col 10 Soft Score
                    "",                                          # col 11 First Failure Tier (backfilled)
                    collapse,                                    # col 12 Collapse
                    "",                                          # col 13 Notes
                ]
                write_row(ws, row_num, data_row, is_pass)

            if not args.dry_run:
                backfill_first_failure(ws, agent_id, model)

    if not args.dry_run:
        wb.save(args.workbook)
        print(f"\nResults saved to {args.workbook}")
    else:
        print("\n[DRY RUN] No API calls made. No file saved.")


if __name__ == "__main__":
    main()
